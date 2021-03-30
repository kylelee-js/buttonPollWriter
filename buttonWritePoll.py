import openpyxl
from openpyxl.utils.cell import coordinate_from_string, start_cell_column_index_from_string


#------------------LOADING---------------------------
print('Start')
filename_target = '기초소방시설설치 설문지.xlsx'

wb = openpyxl.load_workbook(filename_target, data_only=True)
sheet = wb.worksheets[0] 

print('load complete!')

#----------------------------------------------------

start_cell_row = 10
start_cell_col = start_cell_column_index_from_string("J")
count = 0


# set excel to school table

def makeSchoolTable(scnm):

    sheet.cell(start_cell_row ,start_cell_col).value = scnm
    sheet.cell(start_cell_row, start_cell_col + 1).value = "단독주택"
    sheet.cell(start_cell_row + 1, start_cell_col + 1).value = "다세대주택"
    sheet.cell(start_cell_row + 2, start_cell_col + 1).value = "연립주택"
    sheet.cell(start_cell_row + 3, start_cell_col + 1).value = "다중주택"
    start_cell_row += 5  #다음 칸에 새로운 행렬을 만들기 위해 값 변경


#asdasd

def putValueIn():
    sheet.cell(start_cell_row, start_cell_col).value = Fire_ex_arr.count("O")
    sheet.cell(start_cell_row, start_cell_col + 2).value = Fire_ex_arr.count("X")
    sheet.cell(start_cell_row, start_cell_col + 3).value = Fire_sn_arr.count("O")
    sheet.cell(start_cell_row, start_cell_col + 4).value = Fire_sn_arr.count("X")
    
    sheet.cell(start_cell_row, start_cell_col + 6).value = OO
    sheet.cell(start_cell_row, start_cell_col + 7).value = OX
    sheet.cell(start_cell_row, start_cell_col + 8).value = XO
    sheet.cell(start_cell_row, start_cell_col + 9).value = XX

#-----------------------------------------------------

while True:


    # Write school name
    school_name = input("학교 이름 : ")
    start_cell_row = start_cell_row + count
    if school_name == "save":
        print("작업을 완료하고 파일을 저장하겠습니다.")
        break

    # 초기화
    count = 0
    Fire_ex_arr = []
    Fire_sn_arr = []
    OO, OX, XO, XX = 0
    
    makeSchoolTable(school_name)
    

    while True:
        cell = sheet.cell(start_cell_row ,start_cell_col)
        
        House = input("주택종류 : ")
        if House == "exit":
            putValueIn()  # 유무기록을 파일에 작성
            print("다음 작성을 준비합니다.")
            break

        Fire_ex = input("소화기 유뮤 (O, X로 표기) : ")
        Fire_sn = input("경보기 유무 (O, X로 표기) : ")

        Fire_ex_arr.append(Fire_ex)
        Fire_sn_arr.append(Fire_sn)


        if Fire_ex_arr[count] == "O" and Fire_sn_arr[count] == "O":
            OO += 1
        elif Fire_ex_arr[count] == "O" and Fire_sn_arr[count] == "X":
            OX += 1
        elif Fire_ex_arr[count] == "X" and Fire_sn_arr[count] == "O":
            XO += 1
        elif Fire_ex_arr[count] == "X" and Fire_sn_arr[count] == "X":
            XX += 1
        else:
            print("입력이 잘못되었습니다")
        
        count += 1
        print(count)

        
#----------------------------------------------------


print("Start saving")

wb.save("기초소방시설설치 설문지.xlsx")

print("job done!")
